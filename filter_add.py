import pandas as pd
import openpyxl

import data_view

def show(df):
    df1 = openpyxl.load_workbook("svatprj.xlsx")
    page = df1["sss"]
    rows = page.max_row

    print("Show:\n\t1. Department\n\t2. Job\n\t3. Gender\n\t4. Work Day\n\t5. Total Salary\n\t6. Monthly Salary\n\t7. Final Salary\n\t8. Add new\n\tAny. Exit")
    choice = input("\nPlease Choose: ")
    if not data_view.run_df_gui(usercho(choice, df, page, rows, df1)):
        print("Added new employee")


def usercho(choice, df, page, rows, df1):
    match choice:
        case "1":
            print("\t1. General Management\n\t2. Marketing\n\t3. Operations\n\t4. Finance\n\t5. Sales\n\t6. Human Resource\n\t7. Purchase")
            choice2 = input("\nPlease Choose: ")
            if(choice2=="1"):
                cho="General Management"
            elif(choice2=="2"):
                cho="Marketing"
            elif(choice2=="3"):
                cho="Operations"
            elif(choice2=="4"):
                cho="Finance"
            elif(choice2=="5"):
                cho="Sales"
            elif(choice2=="6"):
                cho="Human Resource"
            elif(choice2=="7"):
                cho="Purchase"
            else:
                print("Exiting")
            return df[(df["Dept"] == cho)]
        case "2":
            print("\t1. Engineer\n\t2. Architect\n\t3. Accountant\n\t4. Officer\n\t5. Operator\n\t6. Driver\n\t7. Leader")
            choice2 = input("\nPlease Choose: ")
            if(choice2=="1"):
                cho="Engineer"
            elif(choice2=="2"):
                cho="Architect"
            elif(choice2=="3"):
                cho="Accountant"
            elif(choice2=="4"):
                cho="Officer"
            elif(choice2=="5"):
                cho="Operator"
            elif(choice2=="6"):
                cho="Driver"
            elif(choice2=="7"):
                cho="Leader"
            else:
                print("Exiting")
            return df[(df["Job"] == cho)]
        case "3":
            print("1)Male\n2)Female")
            choice2 = input("\nPlease Choose: ")
            if(choice2=="1"):
                cho="M"
            elif(choice2=="2"):
                cho="F"
            else:
                print("Exiting")
            return df[(df["Gender"] == cho)]
        case "4":
            choice2 = int(input("Show work days higher than(1-30): "))
            return df[(df["Work Day"] > choice2)]
        case "5":
            choice2 = int(input("Show total salary higher than(500-2500): "))
            return df[(df["Total Salary"] > choice2)]
        case "6":
            choice2 = int(input("Show monthly salary higher than: "))
            return df[(df["Monthly Salary"] > choice2)]
        case "7":
            choice2 = int(input("Show final salary higher than: "))
            return df[(df["Final Salary"] > choice2)]
        case "8":
            page.cell(row=rows+1, column = 1, value = rows)
            name = input("Enter Name: ")
            surname = input("Enter Surname: ")
            gender = input("Enter Gender: ")
            dept = input("Enter Department: ")
            job = input("Enter Job: ")
            workday = int(input("Enter Work Day: "))
            total = int(input("Enter Total Salary: "))
            monthly = int(total/30*workday)
            income = int(monthly/20)
            medic = int(monthly/50)
            jobt = int(monthly/33)
            gover = int(monthly/200)
            final = int(monthly-income-medic-jobt-gover)
            page.cell(row=rows+1, column = 2, value = dept)
            page.cell(row=rows+1, column = 3, value = job)
            page.cell(row=rows+1, column = 4, value = gender)
            page.cell(row=rows+1, column = 5, value = name)
            page.cell(row=rows+1, column = 6, value = surname)
            page.cell(row=rows+1, column = 7, value = workday)
            page.cell(row=rows+1, column = 8, value = total)
            page.cell(row=rows+1, column = 9, value = monthly)
            page.cell(row=rows+1, column = 10, value = income)
            page.cell(row=rows+1, column = 11, value = medic)
            page.cell(row=rows+1, column = 12, value = jobt)
            page.cell(row=rows+1, column = 13, value = gover)
            page.cell(row=rows+1, column = 14, value = final)
            df1.save("svatprj.xlsx")
        case default:
            return "Thank you."

            
if __name__ == "__main__":
    main()
